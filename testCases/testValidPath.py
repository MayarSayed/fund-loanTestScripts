from pageObjects.LoanPage import LoanPage
from pageObjects.loginPage import LoginPage
from openpyxl import load_workbook
import pytest
from Utilities.readProperties import ReadConfig
from selenium import webdriver


class Test_Valid_Path:
    # set file path
    filepath="DDT_Data.xlsx"
    # load excel-example.xlsx
    wb=load_workbook(filepath)
    # activate demo.xlsx
    sheet=wb.active

    lenderBalance = sheet.cell(row=2,column=1)
    loanAmount    = sheet.cell(row=2,column=2)
    
    ''' Test Objective : to check valid loan success and appearance of succes messange in case
        of lender balance is greater or equal loan amount
        
        Test Preconditions : 1- Borrower accepted the lender's offer
                             2- Lender decided to fund the loan
    '''
    def testValidLenderBalance(self,setUp):
        self.driver = setUp
        self.driver.get(ReadConfig.getLoanPageURL())
        LP = LoanPage(self.driver)
        
        if(self.lenderBalance >= self.loanAmount):
            LP.clickLoanButton()
            
            if(LP.succesMessageAppears()):
                assert True
            else:
                assert False
        else:
            #Invalid Data for testing
            assert False

        self.driver.close()



