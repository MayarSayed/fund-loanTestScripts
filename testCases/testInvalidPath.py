from pageObjects.LoanPage import LoanPage
from openpyxl import load_workbook
import pytest
from Utilities.readProperties import ReadConfig
from selenium import webdriver


class Test_Valid_Path:
    # set file path
    filepath="DDT_Data.xlsx"
    # load excel-example.xlsx
    wb=load_workbook(filepath)
    # activate demo.xlsx
    sheet=wb.active

    lenderBalance = sheet.cell(row=3,column=1)
    loanAmount    = sheet.cell(row=3,column=2)
    hasBankAccount= sheet.cell(row=3,column=3)
    connectedToBankAcc = sheet.cell(row=3,column=4)
    
    ''' Test Objective : to check loan fail and appearance of error messange in case
        of lender balance is less than loan amount
        
        Test Preconditions : 1- Borrower accepted the lender's offer
                             2- Lender decided to fund the loan
    '''
    def testErrorMessage(self,setUp):
        self.driver = setUp
        self.driver.get(ReadConfig.getLoanPageURL())
        LP = LoanPage(self.driver)
        
        if(self.lenderBalance < self.loanAmount):
            
            if(LP.errorMessageAppears()):
                assert True
            else:
                assert False
        else:
            #Invalid Data for testing
            assert False
            
    ''' Test Objective : to check routing the lender to bank account page in case of loan fail and appearance of error messange in case
        of lender balance is less than loan amount
        
        Test Preconditions : 1- Borrower accepted the lender's offer
                             2- Lender decided to fund the loan
                             3- appearance of error messange in case
                                of lender balance is less than loan amount
    '''
    def testRoutingToBankAccount(self,setUp):
        currentUrl = self.driver.current_url
        
        if currentUrl == ReadConfig.getBankAccountPageURL():
            assert True
        
        else:
            assert False
            
    ''' Test Objective : to check routing the lender to add balance page in case of valid bank account
        
        Test Preconditions : 1- Borrower accepted the lender's offer
                             2- Lender decided to fund the loan
                             3- appearance of error messange in case
                                of lender balance is less than loan amount
                             4- Route to bank account page
                             5- Lender has verified bank account
    '''
    def testRoutingToBalancePage(self,setUp):
        if(self.hasBankAccount):
            currentUrl = self.driver.current_url
            
            if currentUrl == ReadConfig.getAddBalancePageURL():
                assert True
            
            else:
                assert False
                
    ''' Test Objective : to check routing the lender to add bank account page in case of invalid bank account
        
        Test Preconditions : 1- Borrower accepted the lender's offer
                             2- Lender decided to fund the loan
                             3- appearance of error messange in case
                                of lender balance is less than loan amount
                             4- Route to bank account page
                             5- Lender doesn't have verified bank account
    '''
    def testRoutingToAddBankAccount(self,setUp):
        if(not self.hasBankAccount):
            currentUrl = self.driver.current_url
            
            if currentUrl == ReadConfig.getAddBankAccountPageURL():
                assert True
            
            else:
                assert False
                
    ''' Test Objective : to check routing the lender to add balance page in case of connected to bank account succesfully 
                         or stay in the add bank account in case of fail
                         
        Test Preconditions : 1- Borrower accepted the lender's offer
                             2- Lender decided to fund the loan
                             3- appearance of error messange in case
                                of lender balance is less than loan amount
                             4- Route to bank account page
                             5- Lender doesn't have verified bank account
                             6- Lender connected to bank account
    '''
    def testConnectedToBankAccount(self,setUp):
        currentUrl = self.driver.current_url
        
        if(self.connectedToBankAcc):    
            if currentUrl == ReadConfig.getAddBalancePageURL():
                assert True
            
            else:
                assert False
        else:
            if currentUrl == ReadConfig.getAddBankAccountPageURL():
                assert True
            
            else:
                assert False
                
    ''' Test Objective : to check succes message appearence in case lender added the balance successfully
                         & error message in case of failing
                         
        Test Preconditions : 1- Borrower accepted the lender's offer
                             2- Lender decided to fund the loan
                             3- appearance of error messange in case
                                of lender balance is less than loan amount
                             4- Route to bank account page
                             5- Lender doesn't have verified bank account
                             6- Lender connected to bank account
                             7- Lender added balance 
    '''
    def testAddingBalanceS(self,setUp):
        LP = LoanPage(self.driver)
        
        if(self.hasAddedBalance):
            LP.clickLoanButton()
            
            if(LP.succesMessageAppears()):
                assert True
            else:
                assert False
        else:
            if(LP.errorMessageAppears()):
                assert True
            else:
                assert False

        
            
        
    
    



